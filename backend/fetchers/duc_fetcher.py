"""
backend/fetchers/duc_fetcher.py
--------------------------------
Fetches EIA Drilled but Uncompleted (DUC) wells data from:
  https://www.eia.gov/petroleum/drilling/xls/duc-data.xlsx

Also fetches regional rig counts from:
  https://www.eia.gov/petroleum/drilling/xls/dpr-data.xlsx

Writes: backend/data/duc_latest.json

Data is monthly with ~2 month lag.

Key signals:
  - Total DUC count: high = latent supply buffer (bearish), low = supply risk (bullish)
  - DUC drawdown MoM: drawing down = production coming without new drilling (bearish supply signal)
  - Permian DUC: most important single basin
  - Total rigs by region: Permian dominant indicator
"""

import json
import logging
import io
from datetime import datetime, timezone
from pathlib import Path

import requests
import openpyxl

BASE    = Path(__file__).resolve().parents[1]
OUT     = BASE / "data" / "duc_latest.json"

DUC_URL = "https://www.eia.gov/petroleum/drilling/xls/duc-data.xlsx"
DPR_URL = "https://www.eia.gov/petroleum/drilling/xls/dpr-data.xlsx"

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("duc_fetcher")

# ── Regions ───────────────────────────────────────────────────────────────────
REGIONS = ["Anadarko", "Appalachia", "Bakken", "Eagle Ford",
           "Haynesville", "Niobrara", "Permian"]

# DUC thresholds for signal generation
# Historical context: peaked at ~8,874 in June 2020, low ~4,283 in Aug 2022
DUC_HIGH = 7000   # above = ample latent supply buffer (bearish)
DUC_LOW  = 4500   # below = low buffer, supply risk (bullish)


def fetch_excel(url: str) -> openpyxl.Workbook | None:
    """Download and parse Excel file."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        log.info("Downloaded %s — %d bytes", url.split("/")[-1], len(r.content))
        return wb
    except Exception as e:
        log.error("Failed to fetch %s: %s", url, e)
        return None


def _f(v):
    try:
        return float(str(v).replace(",","").strip())
    except:
        return None


def parse_duc_data(wb: openpyxl.Workbook) -> dict:
    """Parse DUC data from webSummary and Data sheets."""
    result = {"by_region": {}, "total": {}, "history": []}

    # ── webSummary: latest 2 months + change ─────────────────────────────────
    if "webSummary" in wb.sheetnames:
        ws    = wb["webSummary"]
        rows  = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]

        # Row 1 (index 0): None, date1, date2, 'change'
        # Row 2+: region, val1, val2, change
        header = rows[0] if rows else []
        dates  = [str(header[1])[:10] if header[1] else None,
                  str(header[2])[:10] if header[2] else None]

        total_latest = 0
        total_prev   = 0

        for row in rows[1:]:
            if not row[0]:
                continue
            region  = str(row[0]).strip()
            if region not in REGIONS:
                continue
            prev    = _f(row[1])
            latest  = _f(row[2])
            change  = _f(row[3])

            result["by_region"][region] = {
                "duc_latest":  latest,
                "duc_prev":    prev,
                "duc_change":  change,
                "period":      dates[1],
                "prev_period": dates[0],
            }
            if latest: total_latest += latest
            if prev:   total_prev   += prev

        result["total"] = {
            "duc_latest":  round(total_latest),
            "duc_prev":    round(total_prev),
            "duc_change":  round(total_latest - total_prev),
            "period":      dates[1],
        }
        log.info("Total DUCs: %d (change: %+d) as of %s",
                 total_latest, total_latest - total_prev, dates[1])

    # ── Data sheet: historical series for Permian ─────────────────────────────
    if "Data" in wb.sheetnames:
        ws   = wb["Data"]
        rows = list(ws.iter_rows(values_only=True))

        # Find column positions
        # Row 3 (index 2): region headers
        # Row 4 (index 3): Drilled, Completed, DUC per region
        region_cols = {}
        if len(rows) > 3:
            header_row  = rows[2]   # region names
            col_row     = rows[3]   # Drilled/Completed/DUC

            # Map regions to their DUC column index
            current_region = None
            for i, val in enumerate(header_row):
                if val and str(val).strip() in REGIONS:
                    current_region = str(val).strip()
                    region_cols[current_region] = {"start_col": i}

            # Find DUC column for each region (3rd column after region start)
            for region, info in region_cols.items():
                start = info["start_col"]
                for j in range(start, min(start + 4, len(col_row))):
                    if col_row[j] and "DUC" in str(col_row[j]):
                        region_cols[region]["duc_col"] = j
                        break

        # Extract last 24 months of total DUC history
        history = []
        for row in rows[4:]:
            if not row[0]:
                continue
            try:
                date = str(row[0])[:10]
                total = 0
                valid = False
                for region, info in region_cols.items():
                    dc = info.get("duc_col")
                    if dc and dc < len(row):
                        v = _f(row[dc])
                        if v is not None:
                            total += v
                            valid = True
                if valid:
                    history.append({"date": date, "total_duc": round(total)})
            except:
                continue

        result["history"] = history[-24:]  # last 24 months

    return result


def parse_dpr_rigs(wb: openpyxl.Workbook) -> dict:
    """Extract latest rig count by region from DPR data."""
    rigs = {}
    total = 0

    for region in REGIONS:
        sheet_name = f"{region} Region"
        if sheet_name not in wb.sheetnames:
            continue
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Data starts at row 3 (index 2), col 1 = rig count
        data_rows = [(r[0], r[1]) for r in rows[2:] if r[0] and r[1]]
        if not data_rows:
            continue

        # Get last 2 rows for MoM
        latest_row = data_rows[-1]
        prev_row   = data_rows[-2] if len(data_rows) > 1 else None

        latest_rigs = _f(latest_row[1])
        prev_rigs   = _f(prev_row[1]) if prev_row else None
        mom         = round(latest_rigs - prev_rigs, 1) if latest_rigs and prev_rigs else None

        rigs[region] = {
            "rigs":   round(latest_rigs, 1) if latest_rigs else None,
            "prev":   round(prev_rigs,   1) if prev_rigs   else None,
            "mom":    mom,
            "period": str(latest_row[0])[:10],
        }

        if latest_rigs:
            total += latest_rigs
            log.info("  %-15s %5.1f rigs (%+.1f MoM)",
                     region, latest_rigs, mom or 0)

    return {"by_region": rigs, "total_rigs": round(total, 1)}


def compute_signal(duc_data: dict, rig_data: dict) -> dict:
    """Generate supply signal from DUC + rig data."""
    total_duc    = duc_data.get("total", {}).get("duc_latest", 0) or 0
    duc_change   = duc_data.get("total", {}).get("duc_change",  0) or 0
    permian_duc  = duc_data.get("by_region", {}).get("Permian", {}).get("duc_latest", 0) or 0
    total_rigs   = rig_data.get("total_rigs", 0) or 0
    permian_rigs = rig_data.get("by_region", {}).get("Permian", {}).get("rigs", 0) or 0

    score    = 0
    reasons  = []

    # DUC level signal
    if total_duc < DUC_LOW:
        score += 2
        reasons.append(f"DUC inventory low ({total_duc:.0f}) — limited latent supply buffer")
    elif total_duc > DUC_HIGH:
        score -= 2
        reasons.append(f"DUC inventory high ({total_duc:.0f}) — large latent supply overhang")

    # DUC trend signal
    if duc_change < -100:
        score -= 1
        reasons.append(f"DUCs drawing down ({duc_change:+.0f} MoM) — completions outpacing drilling")
    elif duc_change > 100:
        score += 1
        reasons.append(f"DUCs building ({duc_change:+.0f} MoM) — drilling outpacing completions")

    # Permian dominance
    if permian_duc and total_duc:
        perm_pct = permian_duc / total_duc * 100
        if perm_pct > 40:
            reasons.append(f"Permian {perm_pct:.0f}% of total DUCs — concentrated supply optionality")

    overall = "BULLISH" if score >= 1 else ("BEARISH" if score <= -1 else "NEUTRAL")

    return {
        "score":          score,
        "overall_signal": overall,
        "reasons":        reasons,
        "key_metrics": {
            "total_duc":    total_duc,
            "duc_change":   duc_change,
            "permian_duc":  permian_duc,
            "total_rigs":   total_rigs,
            "permian_rigs": permian_rigs,
        },
        "interpretation": (
            f"Total DUC inventory: {total_duc:.0f} wells ({duc_change:+.0f} MoM). "
            f"Permian: {permian_duc:.0f} DUCs, {permian_rigs:.0f} rigs. "
            + (reasons[0] if reasons else "DUC inventory within normal range.")
        ),
    }


def run():
    log.info("=" * 60)
    log.info("DUC + DPR FETCHER — EIA Drilling Productivity Report")
    log.info("=" * 60)

    # ── Fetch DUC data ────────────────────────────────────────────────────────
    log.info("Fetching DUC data...")
    duc_wb = fetch_excel(DUC_URL)
    if not duc_wb:
        log.error("Failed to fetch DUC data")
        return {}
    duc_data = parse_duc_data(duc_wb)

    # ── Fetch DPR rig data ────────────────────────────────────────────────────
    log.info("Fetching DPR rig count by region...")
    dpr_wb = fetch_excel(DPR_URL)
    rig_data = parse_dpr_rigs(dpr_wb) if dpr_wb else {"by_region": {}, "total_rigs": 0}

    # ── Signal ────────────────────────────────────────────────────────────────
    signal = compute_signal(duc_data, rig_data)

    output = {
        "fetched_at":  datetime.now(timezone.utc).isoformat(),
        "fetcher":     "duc_fetcher",
        "source":      "EIA Drilling Productivity Report (monthly, ~2 month lag)",
        "duc":         duc_data,
        "rigs":        rig_data,
        "signal":      signal,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w") as f:
        json.dump(output, f, indent=2)

    log.info("─" * 60)
    log.info("Total DUCs:   %d (change: %+d)",
             duc_data.get("total", {}).get("duc_latest", 0),
             duc_data.get("total", {}).get("duc_change", 0))
    log.info("Total Rigs:   %.1f", rig_data.get("total_rigs", 0))
    log.info("Signal:       %s (score=%+d)", signal["overall_signal"], signal["score"])
    log.info("Saved → %s", OUT)
    return output


if __name__ == "__main__":
    run()
